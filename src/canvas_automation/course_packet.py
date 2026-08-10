"""
Logic behind the export-course-packet command.

Two pieces, each split into a pure "build the data" half (unit-testable
with fixtures, no network) and an I/O half that fetches from the local
server and writes files, following the same split used elsewhere in this
engine (see payloads.py).

The gradebook half exists because Canvas has no documented, token-friendly
API endpoint for the gradebook CSV export the web UI produces; the
`/courses/:id/gradebook_csv` route people try requires a browser session,
not a bearer token (confirmed in research/canvas-api-endpoints.md). The
standard workaround, also documented there, is to synthesize an equivalent
table from Enrollments (roster + current/final score) and the bulk
Submissions endpoint (per-assignment score per student).
"""
import csv
import json
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font

from .pdf_tools import merge_pdfs, render_assignment_pdf
from .util import slugify

_HEADER_FONT = Font(name="Calibri", bold=True)
_BODY_FONT = Font(name="Calibri")


def _raw_get(base_url: str, path: str, params: dict) -> list:
    resp = requests.post(
        f"{base_url}/api/raw", json={"method": "GET", "path": path, "params": params}, timeout=60,
    )
    resp.raise_for_status()
    return resp.json()


def _sort_key(sort_by: str):
    if sort_by == "name":
        return lambda a: (a.get("name") or "").lower()
    if sort_by == "position":
        return lambda a: a.get("position") if a.get("position") is not None else 10**9
    return lambda a: a.get("due_at") or "9999"  # due_at (default): undated assignments sort last


def fetch_assignments(base_url: str, course_id: int, only_published: bool, sort_by: str) -> list:
    assignments = _raw_get(base_url, f"/courses/{course_id}/assignments", {"per_page": 100})
    if only_published:
        assignments = [a for a in assignments if a.get("published")]
    assignments.sort(key=_sort_key(sort_by))
    return assignments


def export_assignments_pdf(base_url: str, course_id: int, out_dir: Path, options: dict) -> dict:
    only_published = options.get("only_published", True)
    sort_by = options.get("sort_by", "due_at")

    assignments = fetch_assignments(base_url, course_id, only_published, sort_by)

    json_dir = out_dir / "assignments"
    pdf_dir = out_dir / "assignments_pdf"
    json_dir.mkdir(parents=True, exist_ok=True)
    pdf_dir.mkdir(parents=True, exist_ok=True)

    pdf_paths = []
    for i, assignment in enumerate(assignments, start=1):
        (json_dir / f"{assignment['id']}.json").write_text(json.dumps(assignment, indent=2), encoding="utf-8")
        pdf_path = pdf_dir / f"{i:03d}-{slugify(assignment.get('name'))}.pdf"
        render_assignment_pdf(assignment, pdf_path)
        pdf_paths.append(pdf_path)

    combined_path = out_dir / "assignments_combined.pdf"
    page_count = merge_pdfs(pdf_paths, combined_path) if pdf_paths else 0

    return {
        "assignment_count": len(assignments),
        "combined_pdf": str(combined_path) if pdf_paths else None,
        "page_count": page_count,
    }


def build_gradebook_rows(assignments: list, enrollments: list, submissions_by_student: dict) -> list:
    """
    Pure: assemble the gradebook table (a list of rows, first row is the
    header) from already-fetched data. No network, no filesystem; this is
    what tests/test_engine.py exercises directly with fixture data.
    """
    header = (
        ["Student", "Canvas User ID"]
        + [f"{a.get('name')} ({a.get('points_possible')})" for a in assignments]
        + ["Current Score", "Final Score"]
    )
    rows = [header]
    for enrollment in enrollments:
        user = enrollment.get("user") or {}
        name = user.get("sortable_name") or user.get("name") or f"User {enrollment.get('user_id')}"
        student_scores = submissions_by_student.get(enrollment.get("user_id"), {})
        row = [name, enrollment.get("user_id")]
        for a in assignments:
            row.append(student_scores.get(a["id"], ""))
        grades = enrollment.get("grades") or {}
        row.append(grades.get("current_score", ""))
        row.append(grades.get("final_score", ""))
        rows.append(row)
    return rows


def write_gradebook_csv(rows: list, csv_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


def write_gradebook_xlsx(rows: list, xlsx_path: Path) -> None:
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Gradebook"
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _HEADER_FONT if row_idx == 1 else _BODY_FONT
    for col_idx, header_value in enumerate(rows[0], start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(12, min(40, len(str(header_value)) + 2))
    ws.freeze_panes = "A2"
    wb.save(xlsx_path)


def export_gradebook(base_url: str, course_id: int, out_dir: Path, options: dict) -> dict:
    only_published = options.get("only_published", True)
    also_xlsx = options.get("also_xlsx", True)

    assignments = fetch_assignments(base_url, course_id, only_published, "position")
    assignment_ids = [a["id"] for a in assignments]

    enrollments = _raw_get(
        base_url, f"/courses/{course_id}/enrollments",
        {"type[]": "StudentEnrollment", "state[]": "active", "include[]": "user", "per_page": 100},
    )

    submissions_by_student = {}
    if assignment_ids:
        grouped = _raw_get(
            base_url, f"/courses/{course_id}/students/submissions",
            {"student_ids[]": "all", "assignment_ids[]": assignment_ids, "grouped": "true",
             "include[]": "user", "per_page": 100},
        )
        for entry in grouped:
            submissions_by_student[entry.get("user_id")] = {
                s["assignment_id"]: s.get("score") for s in entry.get("submissions", [])
            }

    rows = build_gradebook_rows(assignments, enrollments, submissions_by_student)

    csv_path = out_dir / "gradebook.csv"
    write_gradebook_csv(rows, csv_path)

    xlsx_path = out_dir / "gradebook.xlsx" if also_xlsx else None
    if xlsx_path:
        write_gradebook_xlsx(rows, xlsx_path)

    return {
        "student_count": len(enrollments),
        "assignment_columns": len(assignments),
        "csv": str(csv_path),
        "xlsx": str(xlsx_path) if xlsx_path else None,
    }
